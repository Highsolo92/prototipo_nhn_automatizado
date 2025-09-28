from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

MASTER_PATH = "data/master.xlsx"
COLUMNS = [
    "id", "created_at", "name", "email", "source",
    "priority", "status", "owner", "notes"
]

def ensure_master():
    os.makedirs("data", exist_ok=True)
    if not os.path.exists(MASTER_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        ws.append(COLUMNS)
        wb.save(MASTER_PATH)
    return MASTER_PATH

def append_lead(lead_dict: dict):
    path = ensure_master()
    wb = load_workbook(path)
    ws = wb["Leads"]
    row = [lead_dict.get(col, "") for col in COLUMNS]
    ws.append(row)
    wb.save(path)

def export_daily(leads: list) -> str:
    os.makedirs("data/reports", exist_ok=True)
    date_str = datetime.now().strftime("%Y%m%d")
    filename = f"data/reports/reporte_{date_str}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws.append(COLUMNS)
    for lead in leads:
        ws.append([
            getattr(lead, col) if hasattr(lead, col) else ""
            for col in COLUMNS
        ])
    wb.save(filename)
    return filename
